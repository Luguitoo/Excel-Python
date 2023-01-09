
from openpyxl import load_workbook
from datetime import datetime

from openpyxl.styles import PatternFill

#Cargamos el archivo
wb = load_workbook('templateprocesos.xlsx')

#Datos
curso = "1"
seccion = "A"
enfasis = "Sociales"
turno = "Noche"
docente = "Docente Test"
disciplina = "Quimica"
etapa = "1"
inf = datetime.now()
# Extraemos la fecha
fecha = datetime.strftime(inf, '%Y/%m/%d')

#Alumnos y trabajos para testear
alumnos = ["Carlos Lugo Zacarias", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Hernesto Mauricio Gimenez Atencio", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez", "Cristiano Ronaldo", "Lionel Messi"
           ,"Carlos Manolo Piris Delgado", "Mathias Ortellado", "Anibal Lopez"]

indicadores = ["Indicador 1 test", "Indicador 2 test", "Indicador 3 test", "Indicador 4 test"]

indicadores2 = ["Indicador 4 test", "Indicador 5 test", "Indicador 6 test", "Indicador 7 test"]

ind1 = ["L", "L", "L", "L", "L", "L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L",
        "L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L"]

ind2 = ["L", "NL", "NL", "L", "L", "L", "L", "NL", "L", "L","L", "L", "NL", "L", "L","L", "NL", "L", "NL", "L","L", "L", "L", "L", "L",
        "L", "NL", "L", "NL", "L","L", "NL", "L", "L", "L","L", "L", "L", "L", "L","L", "NL", "L"]

ind3 = ["L", "NL", "NL", "L", "L", "L", "L", "NL", "L", "L","L", "L", "NL", "L", "L","L", "NL", "L", "NL", "L","L", "L", "L", "L", "L",
        "L", "NL", "L", "NL", "L","L", "NL", "L", "L", "L","L", "L", "L", "L", "L","L", "NL", "L"]

ind4 = ["L", "L", "L", "L", "L", "L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L",
        "L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L", "L", "L","L", "L", "L"]

#print(len(alumnos), len(ind1))
#Para agilizar el proceso
celdas = ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K']

#indicamos la hoja
ws = wb["Hoja1"]
#Insertamos los datos en la cabecera
ws['C4'] = docente
ws['V4'] = disciplina
ws['AP4'] = enfasis
ws['C5'] = etapa
ws['G5'] = 1
ws['T5'] = turno
#Insertamos alumnos
inicio = 12
for x in range(0, len(alumnos)):
    if inicio != 32:
        ws['B{a}'.format(a = str(inicio + 1))] = alumnos[x]

        if inicio == 22:
            #Ejemplo de pintar una celda
            redFill = PatternFill(start_color='cc9cfc',
                                  end_color='cc9cfc',
                                  fill_type='solid')
            ws['B{a}'.format(a=str(inicio + 1))].fill = redFill

        inicio += 1
    else:
        inicio = 42
        ws['B{a}'.format(a=str(inicio + 1))] = alumnos[x]
        inicio += 1
#Insertamos los indicadores
for x in range(0, len(indicadores)):
    ws['{a}11'.format(a= celdas[x])] = indicadores[x]
    ws['{a}11'.format(a=celdas[x + 4])] = indicadores2[x]

#Insertamos los puntajes de cada alumno
inicio = 12
for i in range(0, len(alumnos)):
    print(alumnos[i])
    for x in range(0, len(indicadores)):
        if inicio != 32:
            if x == 0:
                ws['{b}{a}'.format(a = str(inicio + 1), b = celdas[x])] = ind1[i]
            if x == 1:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind2[i]
            if x == 2:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind3[i]
            if x == 3:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind4[i]
        else:
            inicio = 42
            if x == 0:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind1[i]
            if x == 1:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind2[i]
            if x == 2:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind3[i]
            if x == 3:
                ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = ind4[i]
        if x == len(indicadores) - 1:
            inicio += 1
#Cambiar la hoja de nombre
ws.title = curso + " " + seccion + " " + " " + enfasis

#Guarda el archivo
wb.save('{a}_{b}_{c}{d}.xlsx'.format(a = disciplina, b = curso, c = enfasis, d = str(seccion)))